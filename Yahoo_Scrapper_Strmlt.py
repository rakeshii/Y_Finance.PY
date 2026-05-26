import io
import time
import requests
import pandas as pd
import yfinance as yf
import streamlit as st
import plotly.graph_objects as go
from bs4 import BeautifulSoup
from datetime import datetime, date, timedelta
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="FinScraper — Yahoo Finance Exporter",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

PAGE_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Syne:wght@400;600;700;800&display=swap');

/* ==================== BASE STYLES ==================== */
html, body {
    font-family: 'Syne', sans-serif;
}

.stApp {
    background: #0a0e1a;
    color: #e8eaf0;
}

.block-container {
    padding-top: 2rem !important;
    padding-bottom: 2rem !important;
    padding-left: 2rem !important;
    padding-right: 2rem !important;
    max-width: 1400px !important;
}

/* ==================== SIDEBAR ==================== */
section[data-testid="stSidebar"] {
    width: 320px !important;
    min-width: 280px !important;
    max-width: 320px !important;
    background: #0f1626 !important;
    border-right: 1px solid #1e2d4a;
}

/* ==================== HEADER ==================== */
.fin-header {
    padding: 2rem 0 1rem 0;
    border-bottom: 1px solid #1e2d4a;
    margin-bottom: 2rem;
}

.fin-header h1 {
    font-family: 'Syne', sans-serif;
    font-size: clamp(2.2rem, 5vw, 4rem);
    font-weight: 800;
    background: linear-gradient(135deg, #4fc3f7 0%, #00e5ff 50%, #69f0ae 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    margin: 0;
    letter-spacing: -1px;
}

.fin-header p {
    color: #5a7a9a;
    font-family: 'Space Mono', monospace;
    font-size: 0.85rem;
    margin-top: 0.4rem;
    letter-spacing: 0.05em;
}

/* ==================== NATIVE STREAMLIT METRICS STYLING ==================== */
[data-testid="stMetricValue"] {
    font-family: 'Syne', sans-serif;
    font-size: 1.5rem;
    font-weight: 700;
    color: #4fc3f7;
}

[data-testid="stMetricLabel"] {
    font-family: 'Space Mono', monospace;
    font-size: 0.75rem;
    color: #5a7a9a;
    text-transform: uppercase;
    letter-spacing: 0.1em;
}

[data-testid="stMetricDelta"] {
    font-family: 'Syne', sans-serif;
    font-size: 1rem;
}

div[data-testid="metric-container"] {
    background: #111827;
    border: 1px solid #1e2d4a;
    border-radius: 12px;
    padding: 1.2rem;
    position: relative;
    overflow: hidden;
}

div[data-testid="metric-container"]::before {
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 2px;
    background: linear-gradient(90deg, #4fc3f7, #69f0ae);
    z-index: 0;
}

/* ==================== BADGES & LABELS ==================== */
.ticker-badge {
    display: inline-block;
    background: linear-gradient(135deg, #1a2744, #0f1e38);
    border: 1px solid #2a4070;
    border-radius: 8px;
    padding: 0.3rem 0.8rem;
    font-family: 'Space Mono', monospace;
    font-size: 0.85rem;
    color: #4fc3f7;
    margin: 0.2rem;
}

.section-title {
    font-family: 'Syne', sans-serif;
    font-size: 1.1rem;
    font-weight: 700;
    color: #4fc3f7;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin: 2rem 0 1rem 0;
    padding-bottom: 0.5rem;
    border-bottom: 1px solid #1e2d4a;
}

/* ==================== STATUS BOXES ==================== */
.status-box {
    background: #0f1e38;
    border: 1px solid #1e2d4a;
    border-left: 3px solid #4fc3f7;
    border-radius: 8px;
    padding: 0.8rem 1.2rem;
    font-family: 'Space Mono', monospace;
    font-size: 0.8rem;
    color: #5a9acc;
    margin: 0.5rem 0;
}

.status-box.success { 
    border-left-color: #69f0ae; 
    color: #69f0ae; 
}

.status-box.error { 
    border-left-color: #ff5252; 
    color: #ff5252; 
}

.status-box.warning { 
    border-left-color: #ffb300; 
    color: #ffb300; 
}

/* ==================== BUTTONS ==================== */
.stDownloadButton > button {
    background: linear-gradient(135deg, #4fc3f7, #00e5ff) !important;
    color: #0a0e1a !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 700 !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.6rem 2rem !important;
    font-size: 1rem !important;
    letter-spacing: 0.03em !important;
    transition: all 0.2s !important;
    width: 100%;
}

.stDownloadButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 24px rgba(79, 195, 247, 0.35) !important;
}

.stButton > button {
    background: linear-gradient(135deg, #1a2744, #0f1e38) !important;
    color: #4fc3f7 !important;
    border: 1px solid #2a4070 !important;
    border-radius: 10px !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 600 !important;
    width: 100%;
    padding: 0.6rem !important;
    transition: all 0.2s !important;
}

.stButton > button:hover {
    border-color: #4fc3f7 !important;
    box-shadow: 0 4px 12px rgba(79, 195, 247, 0.2) !important;
}

/* ==================== FORM INPUTS ==================== */
.stTextArea textarea {
    background: #111827 !important;
    border: 1px solid #1e2d4a !important;
    border-radius: 10px !important;
    color: #e8eaf0 !important;
    font-family: 'Space Mono', monospace !important;
    font-size: 0.9rem !important;
    transition: all 0.2s !important;
}

.stTextArea textarea:focus {
    border-color: #4fc3f7 !important;
    box-shadow: 0 0 0 2px rgba(79, 195, 247, 0.15) !important;
    outline: none !important;
}

/* ==================== TABS ==================== */
.stTabs [data-baseweb="tab-list"] {
    background: #0f1626;
    border-radius: 10px;
    padding: 4px;
    gap: 4px;
    border: 1px solid #1e2d4a;
}

.stTabs [data-baseweb="tab"] {
    font-family: 'Syne', sans-serif !important;
    color: #5a7a9a !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    transition: all 0.2s !important;
}

.stTabs [data-baseweb="tab"]:hover {
    background: rgba(79, 195, 247, 0.05) !important;
}

.stTabs [aria-selected="true"] {
    background: #1a2744 !important;
    color: #4fc3f7 !important;
}

/* ==================== PROGRESS & DATA ==================== */
.stProgress > div > div {
    background: linear-gradient(90deg, #4fc3f7, #69f0ae) !important;
}

.stDataFrame {
    border: 1px solid #1e2d4a !important;
    border-radius: 10px !important;
    overflow: hidden !important;
}

/* ==================== RESPONSIVE DESIGN ==================== */
@media (max-width: 768px) {
    section[data-testid="stSidebar"] {
        width: 100% !important;
        min-width: 100% !important;
        max-width: 100% !important;
    }
    
    .block-container {
        padding-left: 1rem !important;
        padding-right: 1rem !important;
    }
    
    .fin-header h1 {
        font-size: 2rem;
    }
}

/* ==================== LANDING PAGE CARDS ==================== */
.landing-card {
    background: #111827;
    border: 1px solid #1e2d4a;
    border-radius: 12px;
    padding: 2rem 1rem;
    text-align: center;
    min-height: 220px;
    display: flex;
    flex-direction: column;
    justify-content: center;
    position: relative;
    overflow: hidden;
}

.landing-card::before {
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 2px;
    background: linear-gradient(90deg, #4fc3f7, #69f0ae);
}

.landing-icon {
    font-size: 2.5rem;
    margin-bottom: 1rem;
}

.landing-title {
    font-family: 'Syne', sans-serif;
    font-size: 1rem;
    font-weight: 700;
    color: #e8eaf0;
    margin-bottom: 0.5rem;
}

.landing-desc {
    font-family: 'Space Mono', monospace;
    font-size: 0.72rem;
    color: #5a7a9a;
    line-height: 1.6;
}
</style>
"""
st.markdown(PAGE_CSS, unsafe_allow_html=True)

STATS_SECTION_NAMES = [
    "Valuation Measures", "Fiscal Year", "Profitability",
    "Management Effectiveness", "Income Statement", "Balance Sheet",
    "Cash Flow Statement", "Stock Price History", "Share Statistics",
    "Dividends & Splits",
]
ANALYSIS_SECTION_NAMES = [
    "Earnings Estimate", "Revenue Estimate", "Earnings History",
    "EPS Trend", "EPS Revisions", "Growth Estimates",
]

INTER_TICKER_DELAY = 3.0
RETRY_DELAYS = [5, 15, 30]
SCRAPE_DELAY = 2.0
DURATION_MAP = {
    "1M": 30,
    "3M": 90,
    "6M": 180,
    "1Y": 365,
    "3Y": 365 * 3,
    "5Y": 365 * 5,
    "6Y": 365 * 6,
    "10Y": 365 * 10,
}


def safe_df(value):
    return value if isinstance(value, pd.DataFrame) else pd.DataFrame()


def scrape_yahoo_sections(url, section_names=None, retries=3):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "DNT": "1",
        "Connection": "close",
    }

    for attempt in range(retries):
        try:
            response = requests.get(url, headers=headers, timeout=15)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, "html.parser")

            sections = []
            tables = soup.find_all("table")
            for index, table in enumerate(tables):
                heading = None
                heading_tag = table.find_previous_sibling(["h3", "h2", "strong", "span"])
                if heading_tag:
                    text = heading_tag.get_text(strip=True)
                    if text and len(text) > 2 and text.lower() not in ("--", ""):
                        heading = text

                if not heading and section_names and index < len(section_names):
                    heading = section_names[index]
                if not heading:
                    heading = f"Section {index + 1}"

                rows = []
                for row in table.find_all("tr"):
                    cols = [cell.get_text(strip=True) for cell in row.find_all("td")]
                    if cols:
                        rows.append(cols)

                if rows:
                    sections.append((heading, pd.DataFrame(rows)))

            return sections

        except requests.exceptions.HTTPError as error:
            if error.response is not None and error.response.status_code == 429:
                time.sleep(RETRY_DELAYS[min(attempt, len(RETRY_DELAYS) - 1)])
                continue
            break
        except Exception:
            time.sleep(RETRY_DELAYS[min(attempt, len(RETRY_DELAYS) - 1)])

    return []


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_ticker_data(ticker: str, start_date, end_date):
    for attempt in range(3):
        try:
            stock = yf.Ticker(ticker)
            history_df = safe_df(stock.history(start=start_date, end=end_date))
            if not history_df.empty and isinstance(history_df.index, pd.DatetimeIndex):
                history_df.index = history_df.index.tz_convert(None) if history_df.index.tz else history_df.index
                history_df = history_df.sort_index(ascending=False)

            return {
                "history": history_df,
                "balance_sheet": safe_df(stock.balance_sheet),
                "financials": safe_df(stock.financials),
                "cashflow": safe_df(stock.cashflow),
                "info": stock.info if isinstance(stock.info, dict) else {},
                "error": None,
            }

        except Exception as error:
            message = str(error)
            if any(term in message for term in ["429", "Too Many Requests", "Rate"]):
                time.sleep(RETRY_DELAYS[min(attempt, len(RETRY_DELAYS) - 1)])
                continue
            return {"error": message}

    return {"error": "Rate limited after 3 retries. Please wait a few minutes and try again."}


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_scraped_sections(ticker: str, include_analysis: bool):
    stats_url = f"https://finance.yahoo.com/quote/{ticker}/key-statistics"
    analysis_url = f"https://finance.yahoo.com/quote/{ticker}/analysis"

    stats_sections = scrape_yahoo_sections(stats_url, STATS_SECTION_NAMES)
    time.sleep(SCRAPE_DELAY)
    analysis_sections = []
    if include_analysis:
        analysis_sections = scrape_yahoo_sections(analysis_url, ANALYSIS_SECTION_NAMES)

    return stats_sections, analysis_sections


def format_sheet_for_excel(writer, df, sheet_name):
    sheet_df = df.apply(pd.to_numeric, errors="coerce") / 1000
    sheet_df = sheet_df.round(1)
    sheet_df.to_excel(writer, sheet_name=sheet_name)
    worksheet = writer.sheets[sheet_name]
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                   min_col=1, max_col=worksheet.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0"


def write_sectioned_sheet(writer, sheet_name, sections):
    worksheet = writer.book.create_sheet(sheet_name)
    header_fill = PatternFill("solid", fgColor="0F3460")
    header_font = Font(bold=True, color="4FC3F7", size=11)

    for heading, section_df in sections:
        worksheet.append([heading])
        row_number = worksheet.max_row
        column_count = max(2, section_df.shape[1])
        worksheet.merge_cells(start_row=row_number, start_column=1,
                              end_row=row_number, end_column=column_count)
        cell = worksheet.cell(row=row_number, column=1)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

        for row in section_df.itertuples(index=False):
            worksheet.append(list(row))
        worksheet.append([])

    for column in worksheet.columns:
        max_length = max((len(str(cell.value or "")) for cell in column), default=10)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = min(max_length + 4, 40)


def build_excel(
    ticker,
    history_df,
    balance_sheet_df,
    financials_df,
    cashflow_df,
    stats_sections,
    analysis_sections,
):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        if not history_df.empty:
            history_df.round(2).to_excel(writer, sheet_name="Price History")
        if not balance_sheet_df.empty:
            format_sheet_for_excel(writer, balance_sheet_df, "Balance Sheet")
        if not financials_df.empty:
            format_sheet_for_excel(writer, financials_df, "Income Statement")
        if not cashflow_df.empty:
            format_sheet_for_excel(writer, cashflow_df, "Cashflow")
        if stats_sections:
            write_sectioned_sheet(writer, "Key Statistics", stats_sections)
        if analysis_sections:
            write_sectioned_sheet(writer, "Analysis", analysis_sections)

    buffer.seek(0)
    return buffer


def make_price_chart(history_df, ticker):
    if history_df.empty:
        return None
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=history_df.index,
        y=history_df["Close"],
        mode="lines",
        name="Close",
        line=dict(color="#4fc3f7", width=2),
        fill="tozeroy",
        fillcolor="rgba(79,195,247,0.07)",
    ))
    fig.add_trace(go.Bar(
        x=history_df.index,
        y=history_df["Volume"],
        name="Volume",
        yaxis="y2",
        marker_color="rgba(105,240,174,0.25)",
    ))
    fig.update_layout(
        title=f"{ticker} — Historical Price & Volume",
        paper_bgcolor="#0a0e1a",
        plot_bgcolor="#0f1626",
        font=dict(family="Syne", color="#e8eaf0"),
        xaxis=dict(gridcolor="#1e2d4a", showgrid=True),
        yaxis=dict(gridcolor="#1e2d4a", title="Price"),
        yaxis2=dict(
            overlaying="y",
            side="right",
            title="Volume",
            showgrid=False,
            tickfont=dict(color="#69f0ae"),
        ),
        legend=dict(bgcolor="#0f1626", bordercolor="#1e2d4a"),
        hovermode="x unified",
        height=420,
    )
    return fig


def make_financials_chart(financials_df, ticker):
    if financials_df.empty:
        return None

    rows = ["Total Revenue", "Gross Profit", "Net Income"]
    available_rows = [row for row in rows if row in financials_df.index]
    if not available_rows:
        return None

    chart_df = financials_df.loc[available_rows].T / 1e9
    chart_df.index = [str(value)[:10] for value in chart_df.index]
    colors = ["#4fc3f7", "#69f0ae", "#ff9800"]

    fig = go.Figure()
    for idx, column_name in enumerate(chart_df.columns):
        fig.add_trace(go.Bar(
            name=column_name,
            x=chart_df.index,
            y=chart_df[column_name],
            marker_color=colors[idx % len(colors)],
        ))

    fig.update_layout(
        title=f"{ticker} — Revenue & Profit (USD Billion)",
        barmode="group",
        paper_bgcolor="#0a0e1a",
        plot_bgcolor="#0f1626",
        font=dict(family="Syne", color="#e8eaf0"),
        xaxis=dict(gridcolor="#1e2d4a"),
        yaxis=dict(gridcolor="#1e2d4a", title="USD Billion"),
        legend=dict(bgcolor="#0f1626", bordercolor="#1e2d4a"),
        height=380,
    )
    return fig


def make_balance_chart(balance_df, ticker):
    if balance_df.empty:
        return None

    rows = [
        "Total Assets",
        "Total Liabilities Net Minority Interest",
        "Stockholders Equity",
    ]
    labels = ["Total Assets", "Total Liabilities", "Stockholders Equity"]
    chosen = [(row, label) for row, label in zip(rows, labels) if row in balance_df.index]
    if not chosen:
        return None

    plot_df = pd.DataFrame({
        label: balance_df.loc[row].values / 1e9
        for row, label in chosen
    }, index=[str(value)[:10] for value in balance_df.columns])

    colors = ["#4fc3f7", "#ff5252", "#69f0ae"]
    fig = go.Figure()
    for idx, col_name in enumerate(plot_df.columns):
        fig.add_trace(go.Bar(
            name=col_name,
            x=plot_df.index,
            y=plot_df[col_name],
            marker_color=colors[idx],
        ))

    fig.update_layout(
        title=f"{ticker} — Balance Sheet (USD Billion)",
        barmode="group",
        paper_bgcolor="#0a0e1a",
        plot_bgcolor="#0f1626",
        font=dict(family="Syne", color="#e8eaf0"),
        xaxis=dict(gridcolor="#1e2d4a"),
        yaxis=dict(gridcolor="#1e2d4a", title="USD Billion"),
        legend=dict(bgcolor="#0f1626", bordercolor="#1e2d4a"),
        height=380,
    )
    return fig


def render_sidebar():
    with st.sidebar:
        st.markdown(
            """
            <div style='padding: 1rem 0;'>
                <div style='font-family: Syne, sans-serif; font-size: 1.5rem; font-weight: 800;
                            background: linear-gradient(135deg, #4fc3f7, #69f0ae);
                            -webkit-background-clip: text; -webkit-text-fill-color: transparent;'>
                    📊 FinScraper
                </div>
                <div style='font-family: Space Mono, monospace; font-size: 0.7rem; color: #5a7a9a; margin-top: 0.3rem;'>
                    Yahoo Finance Exporter
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.markdown("---")
        st.markdown("### Enter Tickers")

        ticker_input = st.text_area(
            label="Tickers",
            placeholder="AAPL\nMSFT\nRELIANCE.NS\nCIPLA.NS",
            height=160,
            label_visibility="collapsed",
        )

        include_analysis = st.checkbox("Include Analysis Sheet", value=True)
        include_charts = st.checkbox("Show Charts in App", value=True)
        fetch = st.button("⚡ Fetch & Export", width='stretch')

        st.markdown("---")
        st.markdown(
            """
            <div style='font-family: Space Mono, monospace; font-size: 0.72rem; color: #3a5a7a; line-height: 1.8;'>
                ℹ️ Supports any Yahoo Finance ticker<br>
                🇮🇳 Indian stocks: add .NS (NSE) or .BO (BSE)<br>
                📦 One Excel file per ticker<br>
                ⚡ Data cached for 1 hour
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.markdown("---")
        st.markdown("### Historical Data Duration")

        duration = st.selectbox(
            "Select Duration",
            list(DURATION_MAP.keys()) + ["MAX", "CUSTOM"],
            index=3,
        )

        today = date.today()
        start_date = today - timedelta(days=365)
        end_date = today

        if duration in DURATION_MAP:
            start_date = today - timedelta(days=DURATION_MAP[duration])
        elif duration == "MAX":
            start_date = date(1900, 1, 1)
        elif duration == "CUSTOM":
            start_date = st.date_input("Start Date", value=today - timedelta(days=365))
            end_date = st.date_input("End Date", value=today)

        st.caption(f"Selected Range: {start_date} → {end_date}")

    return {
        "ticker_input": ticker_input,
        "include_analysis": include_analysis,
        "include_charts": include_charts,
        "fetch": fetch,
        "start_date": start_date,
        "end_date": end_date,
    }


def render_header():
    st.markdown(
        """
        <div class='fin-header'>
            <h1>FinScraper</h1>
            <p>YAHOO FINANCE DATA EXPORTER · EXCEL + CHARTS</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_landing():
    st.markdown("---")
    columns = st.columns(3)
    cards = [
        ("📥", "Excel Export", "Price History, Balance Sheet, Income Statement, Cashflow, Statistics & Analysis"),
        ("📈", "Interactive Charts", "Historical Price, Revenue, Profit & Balance Sheet visualization"),
        ("🌍", "Global Market Support", "Supports US, NSE, BSE and all Yahoo Finance supported tickers"),
    ]

    for column, (icon, title, description) in zip(columns, cards):
        with column:
            st.markdown(
                f"""
                <div class='landing-card'>
                    <div class='landing-icon'>{icon}</div>
                    <div class='landing-title'>{title}</div>
                    <div class='landing-desc'>{description}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    st.markdown(
        """
        <div style='margin-top:3rem; text-align:center; font-family:Space Mono, monospace; font-size:0.82rem; color:#3a5a7a;'>
            ← Enter tickers from sidebar and click Fetch & Export
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_fetch_status(tickers):
    badges = " ".join(f"<span class='ticker-badge'>{ticker}</span>" for ticker in tickers)
    st.markdown(f"**Fetching:** {badges}", unsafe_allow_html=True)


def render_metric_cards(metrics):
    """Display metrics using Streamlit's native st.metric() for reliable rendering"""
    cols = st.columns(len(metrics))
    
    for col, (label, value, cls) in zip(cols, metrics):
        with col:
            # For percentage changes (Day Change), display as delta
            if isinstance(value, str) and ('+' in value or '-' in value) and '%' in value:
                # Extract the percentage value for delta
                st.metric(
                    label=label, 
                    value="",  # Empty value since we're showing delta
                    delta=value
                )
            else:
                # Regular metric display
                st.metric(label=label, value=value)


def render_status(message, variant=""):
    variant_class = f" {variant}" if variant else ""
    st.markdown(f"<div class='status-box{variant_class}'>{message}</div>", unsafe_allow_html=True)


render_header()
inputs = render_sidebar()

if not inputs["fetch"]:
    render_landing()
else:
    tickers = [
        ticker.strip().upper()
        for ticker in inputs["ticker_input"].replace(",", "\n").splitlines()
        if ticker.strip()
    ]

    if not tickers:
        st.error("Please enter at least one ticker symbol.")
        st.stop()

    render_fetch_status(tickers)

    for index, ticker in enumerate(tickers):
        st.markdown(f"<div class='section-title'>📌 {ticker}</div>", unsafe_allow_html=True)
        log = st.empty()

        if index > 0:
            render_status(f"⏳ Waiting {INTER_TICKER_DELAY}s before next request to avoid rate limits...", "warning")
            time.sleep(INTER_TICKER_DELAY)

        try:
            render_status(f"⏳ Fetching yfinance data for {ticker}...")
            result = fetch_ticker_data(ticker, inputs["start_date"], inputs["end_date"])

            if result.get("error"):
                render_status(f"❌ Failed for {ticker}: {result['error']}", "error")
                st.markdown("<br>", unsafe_allow_html=True)
                continue

            history_df = result.get("history")
            if not isinstance(history_df, pd.DataFrame):
                history_df = pd.DataFrame()

            balance_sheet_df = result.get("balance_sheet")
            if not isinstance(balance_sheet_df, pd.DataFrame):
                balance_sheet_df = pd.DataFrame()

            financials_df = result.get("financials")
            if not isinstance(financials_df, pd.DataFrame):
                financials_df = pd.DataFrame()

            cashflow_df = result.get("cashflow")
            if not isinstance(cashflow_df, pd.DataFrame):
                cashflow_df = pd.DataFrame()

            info = result.get("info")
            if not isinstance(info, dict):
                info = {}

            render_status(f"⏳ Scraping Key Statistics page for {ticker}...")
            stats_sections, analysis_sections = fetch_scraped_sections(ticker, inputs["include_analysis"])
            render_status(f"✅ Data fetched successfully for {ticker}", "success")

            info_dict = info if isinstance(info, dict) else {}
            change = info_dict.get("regularMarketChangePercent")
            current_price = info_dict.get("currentPrice") or info_dict.get("regularMarketPrice", "N/A")
            market_cap = info_dict.get("marketCap")
            trailing_pe = info_dict.get("trailingPE")
            sector = info_dict.get("sector", "N/A")

            change_str = f"{change*100:+.2f}%" if isinstance(change, (int, float)) else "N/A"
            change_class = "positive" if (change or 0) >= 0 else "negative"
            
            metrics = [
                ("Current Price", current_price, ""),
                ("Day Change", change_str, change_class),
                ("Market Cap", f"${market_cap / 1e9:.2f}B" if isinstance(market_cap, (int, float)) else "N/A", ""),
                (
                    "Trailing P/E",
                    round(trailing_pe, 2) if isinstance(trailing_pe, (int, float)) else (trailing_pe if trailing_pe is not None else "N/A"),
                    "",
                ),
                ("Sector", sector, ""),
            ]
            render_metric_cards(metrics)

            if inputs["include_charts"] and not history_df.empty:
                tab1, tab2, tab3 = st.tabs(["📈 Price & Volume", "💰 Revenue & Profit", "🏦 Balance Sheet"])
                with tab1:
                    price_chart = make_price_chart(history_df, ticker)
                    if price_chart is not None:
                        st.plotly_chart(price_chart, width='stretch')
                    else:
                        st.info("Price chart data not available.")
                with tab2:
                    fig2 = make_financials_chart(financials_df, ticker)
                    if fig2 is not None:
                        st.plotly_chart(fig2, width='stretch')
                    else:
                        st.info("Financials data not available.")
                with tab3:
                    fig3 = make_balance_chart(balance_sheet_df, ticker)
                    if fig3 is not None:
                        st.plotly_chart(fig3, width='stretch')
                    else:
                        st.info("Balance sheet data not available.")

            excel_buffer = build_excel(
                ticker,
                history_df,
                balance_sheet_df,
                financials_df,
                cashflow_df,
                stats_sections,
                analysis_sections,
            )

            filename = (
                f"{ticker.replace('.', '_')}_{inputs['start_date']}_to_{inputs['end_date']}_"
                f"{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            st.download_button(
                label=f"⬇️ Download {ticker} Excel",
                data=excel_buffer,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_{ticker}",
            )

        except Exception as error:
            render_status(f"❌ Failed for {ticker}: {error}", "error")

        st.markdown("<br>", unsafe_allow_html=True)
